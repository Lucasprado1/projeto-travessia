import shutil
from flask import Flask, request, jsonify, send_from_directory, render_template
from openpyxl.utils import FORMULAE
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
from openpyxl.formula.translate import Translator #copy formula
from flask_cors import CORS
from openpyxl.styles import Protection
import datetime
import pyxlsb
import openpyxl
import pandas as pd

DIRETORIOBASES = "./sources/bases/"
DIRETORIO = "./sources/"
DIRETORIO_OPERACOES = "./base_operacoes.xlsx"


global_filename = None

api = Flask(__name__)
CORS(api)

@api.route("/")
def tela_home():
    return render_template('index.html')

@api.route("/arquivos", methods=["GET"])
def lista_arquivos():
    arquivos = []

    for nome_do_arquivo in os.listdir(DIRETORIO):
        endereco_do_arquivo = os.path.join(DIRETORIO, nome_do_arquivo)

        if(os.path.isfile(endereco_do_arquivo)):
            arquivos.append(nome_do_arquivo)

    return jsonify(arquivos)


@api.route("/arquivos/<nome_do_arquivo>",  methods=["GET"])
def get_arquivo(nome_do_arquivo):
    sourceName= nome_do_arquivo.replace("/", "-")
    return send_from_directory(DIRETORIO, sourceName, as_attachment=True)

@api.route("/operacoes",  methods=["GET"])
def get_operacoes():
    return get_column_values(DIRETORIO_OPERACOES, "idOperation")


@api.route("/arquivos", methods=["POST"])
def post_arquivo():
    global global_filename
    uploaded_file = request.files['file']
    if uploaded_file.filename != '':
        nome_do_arquivo = uploaded_file.filename
        uploaded_file.save(os.path.join(DIRETORIOBASES, nome_do_arquivo))
        global_filename = uploaded_file.filename
        return '', 201
    else:
        return 'Nenhum arquivo selecionado.'

@api.route("/uploadModelo", methods=["POST"])
def post_upload_modelo():
    global global_filename
    uploaded_file = request.files['file']
    nome_do_arquivo = request.form['fileName']  # Obtém o nome do arquivo do formulário
    overwriteTemplate = request.form['overwriteModel']  # Obtém o nome do arquivo do formulário

    if overwriteTemplate == "true":
        #criar bkp do modelo atual e salvar o novo com o nome do atual
        extension = nome_do_arquivo.split(".")[-1]
        name = nome_do_arquivo.split(".")[0]
        diretorio, nome_atual = os.path.split(os.path.join(DIRETORIO, nome_do_arquivo))
        novo_caminho = os.path.join(diretorio, f'{name} - bkp.{extension}')
        shutil.move(os.path.join(DIRETORIO, nome_do_arquivo), novo_caminho)

    if uploaded_file.filename != '' and nome_do_arquivo != '':
        nome_do_arquivo_completo = os.path.join(DIRETORIO, nome_do_arquivo)
        uploaded_file.save(nome_do_arquivo_completo)  # Salva o arquivo com o novo nome
        global_filename = nome_do_arquivo
        return '', 201
    else:
        return 'Nenhum arquivo ou nome de arquivo selecionado.', 400

@api.route("/data", methods=["POST"])
def post_data():
    global global_filename
    if request.is_json:
        data = request.get_json()
        print("===============================================================")
        print('Data received:', data)
        if(global_filename): # variável global contendo o filename
            define_operation(global_filename, data) 
        return jsonify({"message": "Dados recebidos com sucesso!"}), 200
    else:
        return jsonify({"error": "Solicitação inválida. Certifique-se de enviar dados JSON."}), 400

@api.route("/checkValues", methods=["POST"])
def check_values():
    data = request.get_json()
    canCreate = verifyOperations(data)
    if canCreate:
        return jsonify({"message": "É possível criar essa operação!"}), 200
    else:
        return jsonify({"error": "Solicitação inválida. Erro ao inserir operação na base de controle."}), 400

# 'Raposo', 'Ibira', 'Atmosfera', 'Barbosa', 'Barreiras', 'FiveSenses', 'GramPoeme',
#  'LotesCia', 'Ommar', 'PatioLusitania', 'EntreSerras', 'Pardini', 'Dpaula'    

# =================================== inicio ==============================
def define_operation(global_filename, data):
    print("define operation", global_filename, data, global_filename[-4:])
    
    if(data["selectedOperation"] == 'Raposo'):
        neo_report_model_raposo(global_filename, data)
    elif(data["selectedOperation"] == 'Ibirapitanga-Terra Luz'):
        if(global_filename[-4:] == "xlsb"):
            convert_xlsb_to_xlsx(f"sources/bases/{global_filename}", "sources/bases/Modelo_ibira_convertido.xlsx")
            neo_report_model_ibira(global_filename, data)  
        else:
            #Somente pega a base recebida e duplica ela com o nome Modelo_ibira_convertido
            shutil.copy(f"sources/bases/{global_filename}", "sources/bases/Modelo_ibira_convertido.xlsx")
            neo_report_model_ibira(global_filename, data) 
    elif(data["selectedOperation"] == 'Atmosfera'):
        neo_report_model_atmosfera(global_filename, data)
    elif(data["selectedOperation"] == 'Five Senses'):
        neo_report_model_fives(global_filename, data)
    elif(data["selectedOperation"] == 'Barreiras'):
        if(global_filename[-4:] == "xlsb"):
            global_filename_convertido = global_filename[:-5] + "_convertido.xlsx"
            convert_xlsb_to_xlsx(f"sources/bases/{global_filename}", f"sources/bases/{global_filename_convertido}")
            
            neo_report_model_barreiras(global_filename_convertido, data)
        else:
            neo_report_model_barreiras(global_filename, data)
    elif(data["selectedOperation"] == 'Lotes e Cia'):
        if(global_filename[-4:] == "xlsb"):
            global_filename_convertido = global_filename[:-5] + "_convertido.xlsx"
            convert_xlsb_to_xlsx(f"sources/bases/{global_filename}", f"sources/bases/{global_filename_convertido}")
            
            neo_report_model_lotesecia(global_filename_convertido, data)
        else:
            neo_report_model_lotesecia(global_filename, data)        
    else:
        if(global_filename[-4:] == "xlsb"):
            global_filename_convertido = global_filename[:-5] + "_convertido.xlsx"
            convert_xlsb_to_xlsx(f"sources/bases/{global_filename}", f"sources/bases/{global_filename_convertido}")
            
            neo_default_pattern(global_filename_convertido, data)
        else:
            neo_default_pattern(global_filename, data)
    

def convert_xlsb_to_xlsx(input_xlsb_filename, output_xlsx_filename):
    # Abrir o arquivo .xlsb
    with pyxlsb.open_workbook(input_xlsb_filename) as wb:
        dest_wb = openpyxl.Workbook()
        for sheetname in wb.sheets:
            dest_ws = dest_wb.create_sheet(title=sheetname)
            for row in wb.get_sheet(sheetname):
                dest_ws.append([item.v for item in row])

    # Salvar o arquivo .xlsx
    dest_wb.save(output_xlsx_filename)

def get_rows_number(working_tab):
    rows = 0
    for max_row, row in enumerate(working_tab, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def grab_formulas(start_line, end_line, working_tab, start_col, end_col):
    for line in range(start_line, end_line):
        for col in range(start_col, end_col):
            celula_origem = working_tab.cell(row=line, column=col)
            celula_destino = working_tab.cell(row=line + 1, column=col)
            celula_destino.value = Translator(celula_origem.value, origin=celula_origem.coordinate).translate_formula(celula_destino.coordinate)

def copy_and_paste_cells(origin_cells, target_cells):
    for origin_row, target_row in zip(origin_cells, target_cells):
        for origin_cell, target_cell in zip(origin_row, target_row):
            target_cell.value = origin_cell.value

def char_to_int(char):
    if 'A' <= char <= 'Z':
        return ord(char) - ord('A') + 1
    else:
        return None  # Caractere inválido

def copy_and_paste_blocks(origin_sheet, target_sheet, origin_range, target_range, rows_number, selected_operation, rows_number_relacao_contrato):
    min_col_origin = char_to_int(origin_range[0][0])
    max_col_origin = char_to_int(origin_range[1][0])
    min_col_target = char_to_int(target_range[0][0])
    max_col_target = char_to_int(target_range[1][0])
    min_row_origin = int(origin_range[0][1])
    min_row_target = int(target_range[0][1])
    rows_number_in_use = int(origin_range[1][1])
    print('operação - tab - origem - destino',selected_operation, target_sheet.title, origin_range, target_range)
    # if(target_sheet.title == 'Relação de Contratos'): #quantidade de linhas em Relação de Contratos
    #     rows_number_in_use = rows_number_relacao_contrato
    # else:
    #     rows_number_in_use = rows_number

    for rows1, rows2 in zip(origin_sheet.iter_rows(min_row=min_row_origin, 
                                                    max_row=rows_number_in_use + 5, 
                                                    min_col=min_col_origin, 
                                                    max_col=max_col_origin),
                            target_sheet.iter_rows(min_row=min_row_target, 
                                                    max_row=rows_number_in_use + 5, 
                                                    min_col=min_col_target, 
                                                    max_col=max_col_target)):
        for cell1, cell2 in zip(rows1, rows2):
            if(selected_operation == 'Raposo' and target_sheet.title == 'Recebíveis' and cell1.column == 1 and cell1.value != None):
                cell_value_str = cell1.value
                cell2.value = int(cell_value_str)
            else:
                cell2.value = cell1.value


def paste_base_contratos(origin_working_tab, max_row_size, target_working_tab, operation):
    array_duplicatas = []
    coluna_quadralote_recebiveis = 1
    if (operation == 'Barreiras'):
        coluna_quadralote_recebiveis = 2
    for row in origin_working_tab.iter_rows(min_row=7, max_row=max_row_size, min_col=coluna_quadralote_recebiveis, max_col=coluna_quadralote_recebiveis):
        for cell in row:
            array_duplicatas.append(cell.value)  
    array_tratado_1 = list(set(array_duplicatas))
    array_tratado = []
    for item in array_tratado_1:
        if item is not None:
            array_tratado.append(item)

    #cola formulas base contratos
    if (operation == 'Raposo'):
        grab_formulas(3, len(array_tratado) + 2, target_working_tab, 3, 8)
    elif (operation == 'Ibirapitanga/Terra Luz'):
        grab_formulas(3, len(array_tratado) + 2, target_working_tab, 3, 9)
    elif (operation == 'Atmosfera'):
        grab_formulas(3, len(array_tratado) + 2, target_working_tab, 3, 8)
    elif (operation == 'Five Senses'):
        grab_formulas(3, len(array_tratado) + 2, target_working_tab, 3, 8)   
    elif (operation == 'Barreiras'):
        grab_formulas(3, len(array_tratado) + 2, target_working_tab, 3, 14)  
    else:
        grab_formulas(3, len(array_tratado) + 2, target_working_tab, 3, 13) 
        
    # cola valores de array_tratado em aba base contratos
    i = 0
    for row in target_working_tab.iter_rows(min_row=3, max_row=len(array_tratado) + 2, min_col=2, max_col=2):
        for cell in row:
            cell.value = array_tratado[i]
        i = i+1

def insert_close_date(operation, closeDate, working_tab):
    #cell = working_tab.cell(row=3, column=12)  # Coluna "L" (12ª coluna)
    cell = working_tab.cell(row=1, column=1)
    iso_date = datetime.datetime.strptime(closeDate, "%Y-%m-%dT%H:%M:%S.%fZ")
    excelCloseDate = iso_date.strftime("%d/%m/%Y")
    cell.value = excelCloseDate

def load_working_tabs(model_report_wb, source_base):
    abas = {}
    abas['aba_destino_base_contrato'] = model_report_wb['Base Contratos']
    abas['aba_origem_relacao_contrato'] = source_base['Relação de Contratos']
    abas['aba_destino_relacao_contrato'] = model_report_wb['Relação de Contratos']
    abas['aba_origem_recebimento'] = source_base['Recebimentos']
    abas['aba_destino_recebimento'] = model_report_wb['Recebimentos']
    abas['aba_origem_recebiveis'] = source_base['Recebíveis']
    abas['aba_destino_recebiveis'] = model_report_wb['Recebíveis']
    
    return abas

def set_ranges(range):
    campos = range.split(':')
    resultados = []
    for parte in campos:
        letra, numero = parte.split('-')
        resultados.append((letra, numero))

    return resultados


def perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_destino, intervalo_recebiveis_destino, intervalo_recebiveis_origem,
                                intervalo_relacao_contrato, intervalo_relacao_contrato_origem, linhas_destino_recebiveis, selected_operation, rows_number, rows_number_relacao_contrato):
    intervalo_recebimento_destino = set_ranges(intervalo_recebimento_destino)
    intervalo_recebimento_origem = set_ranges(intervalo_recebimento_origem)
    intervalo_recebiveis_destino = set_ranges(intervalo_recebiveis_destino)
    intervalo_recebiveis_origem = set_ranges(intervalo_recebiveis_origem)
    intervalo_relacao_contrato = set_ranges(intervalo_relacao_contrato)
    intervalo_relacao_contrato_origem = set_ranges(intervalo_relacao_contrato_origem)
    print('intervalo rela contratos', intervalo_relacao_contrato_origem)

    #cola recebimento 
    copy_and_paste_blocks(tabs['aba_origem_recebimento'], tabs['aba_destino_recebimento'],
                          intervalo_recebimento_origem, intervalo_recebimento_destino, rows_number, selected_operation, rows_number_relacao_contrato)
    print("colou reebimento")
    #cola recebiveis
    # copy_and_paste_cells(tabs['aba_origem_recebiveis'][intervalo_recebiveis_origem], tabs['aba_destino_recebiveis'][intervalo_recebiveis_destino])

    copy_and_paste_blocks(tabs['aba_origem_recebiveis'], tabs['aba_destino_recebiveis'],
                          intervalo_recebiveis_origem, intervalo_recebiveis_destino, rows_number + 7, selected_operation, rows_number_relacao_contrato)
    print("colou recebiveis", tabs['aba_destino_recebiveis'])
   
    #cola relação contrato
    # copy_and_paste_cells(tabs['aba_origem_relacao_contrato'][intervalo_relacao_contrato], tabs['aba_destino_relacao_contrato'][intervalo_relacao_contrato])

    copy_and_paste_blocks(tabs['aba_origem_relacao_contrato'], tabs['aba_destino_relacao_contrato'],
                          intervalo_relacao_contrato_origem, intervalo_relacao_contrato, rows_number, selected_operation, rows_number_relacao_contrato)
    print("colou relac contratos")
    # pega array com duplicatas em recebiveis
    paste_base_contratos(tabs['aba_destino_recebiveis'], rows_number, tabs['aba_destino_base_contrato'], selected_operation )
    print("colou base contratos")

def neo_report_model_ibira(base_filename, data):
    print('data', base_filename, data)
    model_report_wb = load_workbook("sources/modelo - Ibira.xlsx")
    source_base = load_workbook(f"sources/bases/Modelo_ibira_convertido.xlsx")

    tabs = load_working_tabs(model_report_wb, source_base)

    aba_origem_recebimento = tabs['aba_origem_recebimento'] # inicialização necessária para definir o 'intervalo_recebimento_origem'
    aba_origem_recebiveis = tabs['aba_origem_recebiveis']   #                             ''
    aba_origem_relacao_contrato= tabs['aba_origem_relacao_contrato']
    
    linhas_destino_recebiveis = 0
    # inputa data de fechamento
    insert_close_date(data["selectedOperation"], data["selectedDate"], tabs['aba_destino_recebiveis'])
    #joga formula recebimento 
    #(start_line , x, x, start_col, end_col)
    grab_formulas(2, get_rows_number(tabs['aba_origem_recebimento']),tabs['aba_destino_recebimento'], 20, 27)
    #joga formula recebiveis
    grab_formulas(7, get_rows_number(tabs['aba_origem_recebiveis']) + 5, tabs['aba_destino_recebiveis'], 14, 20)


    intervalo_recebimento_origem = f'A-2:S-{get_rows_number(aba_origem_recebimento)}'
    intervalo_recebiveis_destino = f'A-7:M-{get_rows_number(aba_origem_recebiveis) + 7}'
    intervalo_recebiveis_origem = f'A-2:M-{get_rows_number(aba_origem_recebiveis)}'
    intervalo_relacao_contrato = f'A-2:L-{get_rows_number(aba_origem_relacao_contrato)}'
    intervalo_relacao_contrato_origem = f'A-2:L-{get_rows_number(aba_origem_relacao_contrato)}'

    perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_origem, intervalo_recebiveis_destino, intervalo_recebiveis_origem, intervalo_relacao_contrato, intervalo_relacao_contrato_origem, linhas_destino_recebiveis, data["selectedOperation"], get_rows_number(aba_origem_recebiveis), get_rows_number(aba_origem_relacao_contrato))
    sourceName= data["selectedOperation"].replace("/", "-")
    model_report_wb.save(f"sources/EDT-{sourceName}-{data['userEmail']}.xlsx")

def neo_report_model_raposo(base_filename, data):
    print(data["userEmail"], 'data')
    # model_report_wb = load_workbook("sources/Modelo Relatório - NEO - RAPOSO.xlsx")
    model_report_wb = load_workbook("sources/modelo - Raposo.xlsx")
    source_base = load_workbook(f"sources/bases/{base_filename}")
    

    tabs = load_working_tabs(model_report_wb, source_base)

    aba_origem_recebimento = tabs['aba_origem_recebimento'] # inicialização necessária para definir o 'intervalo_recebimento_origem'
    aba_origem_recebiveis = tabs['aba_origem_recebiveis']   #                             ''
    aba_origem_relacao_contratos = tabs['aba_origem_relacao_contrato']   #  ''

    linhas_destino_recebiveis = 0
    # inputa data de fechamento
    insert_close_date(data["selectedOperation"], data["selectedDate"], tabs['aba_destino_recebiveis'])
    #joga formula recebimento 
    grab_formulas(2, get_rows_number(tabs['aba_origem_recebimento']),tabs['aba_destino_recebimento'], 19, 26)
    #joga formula recebiveis
    grab_formulas(7, get_rows_number(tabs['aba_origem_recebiveis']) + 5, tabs['aba_destino_recebiveis'], 13,19)

    intervalo_recebimento_origem = f'A-2:R-{get_rows_number(aba_origem_recebimento)}'
    intervalo_recebiveis_destino = f'A-7:L-{get_rows_number(aba_origem_recebiveis) + 7}'
    intervalo_recebiveis_origem = f'A-2:L-{get_rows_number(aba_origem_recebiveis)}'
    intervalo_relacao_contrato = f'A-2:K-{get_rows_number(aba_origem_relacao_contratos)}'

    perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_origem, intervalo_recebiveis_destino, intervalo_recebiveis_origem, intervalo_relacao_contrato, intervalo_relacao_contrato, linhas_destino_recebiveis, data["selectedOperation"], get_rows_number(aba_origem_recebiveis), get_rows_number(aba_origem_recebiveis))

    # f"sources/bases/{data["userEmail"]}"
    model_report_wb.save(f"sources/EDT-{data['selectedOperation']}-{data['userEmail']}.xlsx")
    

def neo_default_pattern(base_filename, data):
    print(data["userEmail"], 'data')
    model_report_wb = load_workbook(f"sources/modelo - {data['selectedOperation']}.xlsx")
    source_base = load_workbook(f"sources/bases/{base_filename}")
    

    tabs = load_working_tabs(model_report_wb, source_base)

    aba_origem_recebimento = tabs['aba_origem_recebimento'] # inicialização necessária para definir o 'intervalo_recebimento_origem'
    aba_origem_recebiveis = tabs['aba_origem_recebiveis']   #     
    aba_origem_relacao_contratos = tabs['aba_origem_relacao_contrato']   #  ''

    linhas_destino_recebiveis = 0
    # inputa data de fechamento
    insert_close_date(data["selectedOperation"], data["selectedDate"], tabs['aba_destino_recebiveis'])
    #joga formula recebimento 
    grab_formulas(2, get_rows_number(tabs['aba_origem_recebimento']),tabs['aba_destino_recebimento'], 19, 27)
    #joga formula recebiveis
    grab_formulas(7, get_rows_number(tabs['aba_origem_recebiveis']) + 5, tabs['aba_destino_recebiveis'], 13,20)
    
    intervalo_recebimento_origem = f'A-2:R-{get_rows_number(aba_origem_recebimento)}'
    intervalo_recebiveis_destino = f'A-7:L-{get_rows_number(aba_origem_recebiveis) + 7}'
    intervalo_recebiveis_origem = f'A-2:L-{get_rows_number(aba_origem_recebiveis)}'
    intervalo_relacao_contrato = f'A-2:K-{get_rows_number(aba_origem_relacao_contratos)}'

    perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_origem, intervalo_recebiveis_destino, intervalo_recebiveis_origem, intervalo_relacao_contrato, intervalo_relacao_contrato, linhas_destino_recebiveis, data["selectedOperation"], get_rows_number(aba_origem_recebiveis), get_rows_number(aba_origem_recebiveis))

    # f"sources/bases/{data["userEmail"]}"
    sourceName= data["selectedOperation"].replace("/", "-")
    model_report_wb.save(f"sources/EDT-{sourceName}-{data['userEmail']}.xlsx")
        

def neo_report_model_atmosfera(base_filename, data):
    print('data')
    # model_report_wb = load_workbook("sources/Modelo Relatório - NEO - RAPOSO.xlsx")
    model_report_wb = load_workbook("sources/modelo - Atmosfera.xlsx")
    source_base = load_workbook(f"sources/bases/{base_filename}")
    

    tabs = load_working_tabs(model_report_wb, source_base)

    aba_origem_recebimento = tabs['aba_origem_recebimento'] # inicialização necessária para definir o 'intervalo_recebimento_origem'
    aba_origem_recebiveis = tabs['aba_origem_recebiveis']   #                             ''
    aba_origem_relacao_contratos = tabs['aba_origem_relacao_contrato']   #  ''

    linhas_destino_recebiveis = 0
    # inputa data de fechamento
    insert_close_date(data["selectedOperation"], data["selectedDate"], tabs['aba_destino_recebiveis'])
    #joga formula recebimento 
    grab_formulas(2, get_rows_number(tabs['aba_origem_recebimento']),tabs['aba_destino_recebimento'], 19, 26)
    #joga formula recebiveis
    grab_formulas(7, get_rows_number(tabs['aba_origem_recebiveis']) + 5, tabs['aba_destino_recebiveis'], 13,19)

    intervalo_recebimento_origem = f'A-2:R-{get_rows_number(aba_origem_recebimento)}'
    intervalo_recebiveis_destino = f'A-7:L-{get_rows_number(aba_origem_recebiveis) + 7}'
    intervalo_recebiveis_origem = f'A-2:L-{get_rows_number(aba_origem_recebiveis)}'
    intervalo_relacao_contrato = f'A-2:K-{get_rows_number(aba_origem_relacao_contratos)}'

    perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_origem, intervalo_recebiveis_destino, intervalo_recebiveis_origem, intervalo_relacao_contrato, intervalo_relacao_contrato, linhas_destino_recebiveis, data["selectedOperation"], get_rows_number(aba_origem_recebiveis), get_rows_number(aba_origem_recebiveis))

    model_report_wb.save(f"sources/EDT-{data['selectedOperation']}-{data['userEmail']}.xlsx")


def neo_report_model_fives(base_filename, data):
    print('data', data)
    # model_report_wb = load_workbook("sources/Modelo Relatório - NEO - RAPOSO.xlsx")
    model_report_wb = load_workbook("sources/modelo - fiveSenses.xlsx")
    source_base = load_workbook(f"sources/bases/{base_filename}")
    
    print("entrou função")
    tabs = load_working_tabs(model_report_wb, source_base)

    aba_origem_recebimento = tabs['aba_origem_recebimento'] # inicialização necessária para definir o 'intervalo_recebimento_origem'
    aba_origem_recebiveis = tabs['aba_origem_recebiveis']   #                             ''
    aba_origem_relacao_contratos = tabs['aba_origem_relacao_contrato']   #  ''

    linhas_destino_recebiveis = 0
    # inputa data de fechamento
    
    insert_close_date(data["selectedOperation"], data["selectedDate"], tabs['aba_destino_recebiveis'])
    #joga formula recebimento 
    grab_formulas(2, get_rows_number(tabs['aba_origem_recebimento']),tabs['aba_destino_recebimento'], 19, 26)
    #joga formula recebiveis
    grab_formulas(get_rows_number(tabs['aba_destino_recebiveis']), get_rows_number(tabs['aba_origem_recebiveis']) + 5, tabs['aba_destino_recebiveis'], 13,23)
    intervalo_recebimento_origem = f'A-2:R-{get_rows_number(aba_origem_recebimento)}'
    intervalo_recebiveis_destino = f'A-7:L-{get_rows_number(aba_origem_recebiveis) + 7}'
    intervalo_recebiveis_origem = f'A-2:L-{get_rows_number(aba_origem_recebiveis)}'
    intervalo_relacao_contrato = f'A-2:K-{get_rows_number(aba_origem_relacao_contratos)}'
    perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_origem, intervalo_recebiveis_destino, intervalo_recebiveis_origem, intervalo_relacao_contrato, intervalo_relacao_contrato, linhas_destino_recebiveis, data["selectedOperation"], get_rows_number(aba_origem_recebiveis), get_rows_number(aba_origem_recebiveis))
    
    model_report_wb.save(f"sources/EDT-{data['selectedOperation']}-{data['userEmail']}.xlsx")


def neo_report_model_barreiras(base_filename, data):
    print(data["userEmail"], 'data')
    model_report_wb = load_workbook(f"sources/modelo - {data['selectedOperation']}.xlsx")
    source_base = load_workbook(f"sources/bases/{base_filename}")
    

    tabs = load_working_tabs(model_report_wb, source_base)

    aba_origem_recebimento = tabs['aba_origem_recebimento'] # inicialização necessária para definir o 'intervalo_recebimento_origem'
    aba_origem_recebiveis = tabs['aba_origem_recebiveis']   #     
    aba_origem_relacao_contratos = tabs['aba_origem_relacao_contrato']   #  ''

    linhas_destino_recebiveis = 0
    # inputa data de fechamento
    insert_close_date(data["selectedOperation"], data["selectedDate"], tabs['aba_destino_recebiveis'])
    #joga formula recebimento 
    grab_formulas(2, get_rows_number(tabs['aba_origem_recebimento']),tabs['aba_destino_recebimento'], 20, 28)
    #joga formula recebiveis
    grab_formulas(7, get_rows_number(tabs['aba_origem_recebiveis']) + 5, tabs['aba_destino_recebiveis'], 14,21)
    
    intervalo_recebimento_origem = f'A-2:R-{get_rows_number(aba_origem_recebimento)}' # intervalo recebimento origem != recebimento destino apenas em barreiras
    intervalo_recebimento_destino = f'B-2:S-{get_rows_number(aba_origem_recebimento)}'
    intervalo_recebiveis_destino = f'B-7:M-{get_rows_number(aba_origem_recebiveis) + 7}'
    intervalo_recebiveis_origem = f'A-2:L-{get_rows_number(aba_origem_recebiveis)}'
    intervalo_relacao_contrato = f'B-2:L-{get_rows_number(aba_origem_relacao_contratos)}'
    intervalo_relacao_contrato_origem = f'A-2:K-{get_rows_number(aba_origem_relacao_contratos)}'

    perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_destino ,intervalo_recebiveis_destino, intervalo_recebiveis_origem, intervalo_relacao_contrato, intervalo_relacao_contrato_origem, linhas_destino_recebiveis, data["selectedOperation"], get_rows_number(aba_origem_recebiveis), get_rows_number(aba_origem_recebiveis))

    # f"sources/bases/{data["userEmail"]}"
    sourceName= data["selectedOperation"].replace("/", "-")
    model_report_wb.save(f"sources/EDT-{sourceName}-{data['userEmail']}.xlsx")
        

def neo_report_model_lotesecia(base_filename, data):
    print(data["userEmail"], 'data')
    model_report_wb = load_workbook(f"sources/modelo - {data['selectedOperation']}.xlsx")
    source_base = load_workbook(f"sources/bases/{base_filename}")
    

    tabs = load_working_tabs(model_report_wb, source_base)

    aba_origem_recebimento = tabs['aba_origem_recebimento'] # inicialização necessária para definir o 'intervalo_recebimento_origem'
    aba_origem_recebiveis = tabs['aba_origem_recebiveis']   #     
    aba_origem_relacao_contratos = tabs['aba_origem_relacao_contrato']   #  ''

    linhas_destino_recebiveis = 0
    # inputa data de fechamento
    insert_close_date(data["selectedOperation"], data["selectedDate"], tabs['aba_destino_recebiveis'])
    #joga formula recebimento 
    grab_formulas(2, get_rows_number(tabs['aba_origem_recebimento']),tabs['aba_destino_recebimento'], 19, 27)
    #joga formula recebiveis
    grab_formulas(7, get_rows_number(tabs['aba_origem_recebiveis']) + 5, tabs['aba_destino_recebiveis'], 13,21)
    
    intervalo_recebimento_origem = f'A-2:R-{get_rows_number(aba_origem_recebimento)}'
    intervalo_recebiveis_destino = f'A-7:L-{get_rows_number(aba_origem_recebiveis) + 7}'
    intervalo_recebiveis_origem = f'A-2:L-{get_rows_number(aba_origem_recebiveis)}'
    intervalo_relacao_contrato = f'A-2:K-{get_rows_number(aba_origem_relacao_contratos)}'

    perform_data_copy_and_paste(tabs, intervalo_recebimento_origem, intervalo_recebimento_origem, intervalo_recebiveis_destino, intervalo_recebiveis_origem, intervalo_relacao_contrato, intervalo_relacao_contrato, linhas_destino_recebiveis, data["selectedOperation"], get_rows_number(aba_origem_recebiveis), get_rows_number(aba_origem_recebiveis))

    # f"sources/bases/{data["userEmail"]}"
    sourceName= data["selectedOperation"].replace("/", "-")
    model_report_wb.save(f"sources/EDT-{sourceName}-{data['userEmail']}.xlsx")
        
                
def get_column_values(caminho_arquivo, nome_coluna):
    df = pd.read_excel(caminho_arquivo)
    
    if nome_coluna in df.columns:
        valores_coluna = df[nome_coluna].tolist()
        return valores_coluna
    else:
        return []

def verifyOperations(data):
    try:
        # Abre o arquivo Excel
        df = pd.read_excel(DIRETORIO_OPERACOES)
        # Verifica a existência dos valores nas colunas correspondentes
        if(not data['overwriteModel']):
            opId_exists = 'idOperation' in df and data['idOperation'] in df['idOperation'].values
            excelName_exists = 'excelName' in df and data['excelName'] in df['excelName'].values 
            if not (opId_exists or excelName_exists):
                # Os valores não existem no Excel, então vamos inserir uma nova linha
                row = data.copy()
                row.pop('overwriteModel')
                nova_linha = pd.DataFrame(row, index=[0])  # Cria um novo DataFrame com os valores
                df = pd.concat([df, nova_linha], ignore_index=True)  # Concatena o novo DataFrame ao original

            df.to_excel(DIRETORIO_OPERACOES, index=False)  # Salva o DataFrame de volta no Excel
        return True
    except FileNotFoundError:
        return False  # Se o arquivo não existe, retornar True

if __name__ == '__main__':
    api.run(host='127.0.0.1', port=5000, debug=True, threaded=True)   

    